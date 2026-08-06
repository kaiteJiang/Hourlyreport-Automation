from datetime import date, datetime

import json
import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from modules.baidu_report_api import (
    SEARCH_WORD_COLUMNS,
    SEARCH_WORD_REPORT_TYPE,
    _build_search_word_payload,
    aggregate_search_word_rows,
)
from modules.excel_writer import (
    _scan_word_share_headers,
    _word_share_cell_date_matches,
    write_word_share_data,
)
from modules.kst_daily_aggregation import aggregate_word_class_conversations
from modules.project_config import get_word_share_path
from gui.command_builder import build_word_class_command


def _word_share_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "词类占比日数据"
    ws.append(["", "", "", "", "", ""])
    ws.append(["日期", "点击", "消费", "有效对话", "有效转潜", "到诊"])
    for offset in range(365):
        ws.cell(row=3 + offset, column=1, value=46023 + offset)
    return wb


def _word_share_config(tmp_path):
    return {
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "excel_path": str(tmp_path / "竞价数据.xlsx"),
        "word_share_sheet_name": "词类占比日数据",
    }


def _word_share_merged(date_text: str, **metrics):
    payload = {
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "date": date_text,
        "metrics": {
            "点击": metrics.get("点击", 123),
            "消费": metrics.get("消费", 45.67),
            "有效对话": metrics.get("有效对话", 5),
            "有效转潜": metrics.get("有效转潜", 2),
            "到诊": "",
        },
    }
    return payload


class _NullLogger:
    def info(self, *args, **kwargs):
        pass

    def warning(self, *args, **kwargs):
        pass

    def error(self, *args, **kwargs):
        pass


# ---- 百度搜索词聚合 ----

def test_aggregate_search_word_rows_filters_and_sums():
    rows = [
        {"queryWord": "银屑病怎么治", "click": 3, "cost": 12.5, "impression": 100},
        {"queryWord": "牛皮癣症状", "click": 2, "cost": 8.0, "impression": 80},
        {"queryWord": "痘痘怎么去", "click": 9, "cost": 30.0, "impression": 300},
        {"queryWord": "治疗银屑病医院", "click": 1, "cost": 4.25, "impression": 40},
    ]
    result = aggregate_search_word_rows(rows)

    assert result["matched_rows"] == 3
    assert result["click"] == 6
    assert result["cost"] == 24.75
    assert result["impression"] == 220
    assert result["keyword_counts"] == {"银屑病": 2, "牛皮癣": 1}
    assert result["errors"] == []


def test_aggregate_search_word_rows_rejects_bad_numbers():
    rows = [
        {"queryWord": "银屑病", "click": "abc", "cost": 1.0, "impression": 1},
    ]
    result = aggregate_search_word_rows(rows)
    assert result["matched_rows"] == 1
    assert result["click"] == 0
    assert result["errors"]


def test_build_search_word_payload_uses_search_word_report_type():
    payload = _build_search_word_payload(
        "user", "token", [123, 456], "2026-08-04", start_row=1000, row_count=500
    )
    body = payload["body"]
    assert body["reportType"] == SEARCH_WORD_REPORT_TYPE
    assert "queryWord" in body["columns"]
    assert body["columns"] == SEARCH_WORD_COLUMNS
    assert body["startRow"] == 1000
    assert body["rowCount"] == 500
    assert body["needSum"] is False


# ---- KST 词类对话聚合 ----

def test_aggregate_word_class_conversations_counts_valid_and_lead():
    conversations = [
        {"keyword": "银屑病怎么治", "bid_word": "", "tags": ["有效-三句"]},
        {"keyword": "牛皮癣", "bid_word": "", "tags": ["转潜-有效"]},
        {"keyword": "银屑病", "bid_word": "", "tags": ["有效-一般"]},
        {"keyword": "痘痘", "bid_word": "", "tags": ["有效-三句"]},
        {"keyword": "牛皮癣", "bid_word": "", "tags": ["无效"]},
        {"keyword": "银屑病", "bid_word": "", "tags": []},
    ]
    result = aggregate_word_class_conversations(conversations)

    assert result["matched_conversations"] == 5
    assert result["counts"]["总对话"] == 5
    assert result["counts"]["有效对话"] == 3  # 有效-三句 + 转潜-有效 + 有效-一般
    assert result["counts"]["有效转潜"] == 1  # 仅转潜-有效
    assert result["keyword_counts"] == {"银屑病": 3, "牛皮癣": 2}


def test_aggregate_word_class_conversations_does_not_match_bid_word():
    conversations = [
        {"keyword": "", "bid_word": "牛皮癣", "tags": ["有效-三句"]},
        {"keyword": "银屑病", "bid_word": "无关", "tags": ["转潜-有效"]},
    ]
    result = aggregate_word_class_conversations(conversations)
    # 只按搜索关键词匹配,竞价词命中不算
    assert result["matched_conversations"] == 1
    assert result["counts"]["有效对话"] == 1
    assert result["counts"]["有效转潜"] == 1


# ---- 词类占比 Excel 日期匹配 ----

def test_word_share_cell_date_matches_serial():
    assert _word_share_cell_date_matches(46023, date(2026, 1, 1)) is True
    assert _word_share_cell_date_matches(46023.0, date(2026, 1, 1)) is True
    assert _word_share_cell_date_matches(46024, date(2026, 1, 1)) is False
    assert _word_share_cell_date_matches(datetime(2026, 1, 1, 9, 0), date(2026, 1, 1)) is True
    assert _word_share_cell_date_matches("2026-01-01", date(2026, 1, 1)) is True


def test_scan_word_share_headers():
    wb = _word_share_workbook()
    structure = _scan_word_share_headers(wb["词类占比日数据"])
    columns = structure["columns"]
    assert structure["header_row"] == 2
    assert columns["日期"] == 1
    assert columns["点击"] == 2
    assert columns["消费"] == 3
    assert columns["有效对话"] == 4
    assert columns["有效转潜"] == 5
    assert columns["到诊"] == 6


def test_scan_word_share_headers_missing_field():
    wb = Workbook()
    ws = wb.active
    ws.title = "词类占比日数据"
    ws.append(["日期", "点击", "消费", "有效对话"])
    columns = _scan_word_share_headers(ws)["columns"]
    assert "有效转潜" not in columns


# ---- 词类占比 Excel 写入 ----

def test_write_word_share_data_writes_and_verifies(tmp_path):
    word_share_path = tmp_path / "【昆明牛】2026词类占比数据.xlsx"
    wb = _word_share_workbook()
    wb.save(word_share_path)

    merged_path = tmp_path / "reports" / "word_share_data.json"
    merged_path.parent.mkdir(parents=True, exist_ok=True)
    merged_path.write_text(
        json.dumps(_word_share_merged("2026-01-01"), ensure_ascii=False),
        encoding="utf-8",
    )

    config = _word_share_config(tmp_path)
    report = write_word_share_data(config, tmp_path, _NullLogger(), target_date="2026-01-01")

    assert not report["errors"]
    assert report["self_check"]["backup_created"]
    assert report["self_check"]["verification_passed"]
    assert report["backup_path"]
    assert Path(report["backup_path"]).exists()

    wb = load_workbook(word_share_path)
    ws = wb["词类占比日数据"]
    assert ws.cell(row=3, column=1).value == 46023  # 2026-01-01 行
    assert ws.cell(row=3, column=2).value == 123  # 点击
    assert ws.cell(row=3, column=3).value == 45.67  # 消费
    assert ws.cell(row=3, column=4).value == 5  # 有效对话
    assert ws.cell(row=3, column=5).value == 2  # 有效转潜
    wb.close()


def test_write_word_share_data_keeps_visit_field_untouched(tmp_path):
    word_share_path = tmp_path / "【昆明牛】2026词类占比数据.xlsx"
    wb = _word_share_workbook()
    wb["词类占比日数据"].cell(row=3, column=6, value="已有到诊值")
    wb.save(word_share_path)

    merged_path = tmp_path / "reports" / "word_share_data.json"
    merged_path.parent.mkdir(parents=True, exist_ok=True)
    merged_path.write_text(
        json.dumps(_word_share_merged("2026-01-01"), ensure_ascii=False),
        encoding="utf-8",
    )

    report = write_word_share_data(_word_share_config(tmp_path), tmp_path, _NullLogger(), target_date="2026-01-01")
    assert not report["errors"]

    wb = load_workbook(word_share_path)
    assert wb["词类占比日数据"].cell(row=3, column=6).value == "已有到诊值"
    wb.close()


def test_write_word_share_data_date_row_missing(tmp_path):
    word_share_path = tmp_path / "【昆明牛】2025词类占比数据.xlsx"
    wb = _word_share_workbook()  # 表内是 2026 序列号
    wb.save(word_share_path)

    merged_path = tmp_path / "reports" / "word_share_data.json"
    merged_path.parent.mkdir(parents=True, exist_ok=True)
    merged_path.write_text(
        json.dumps(_word_share_merged("2025-12-31"), ensure_ascii=False),
        encoding="utf-8",
    )

    report = write_word_share_data(_word_share_config(tmp_path), tmp_path, _NullLogger(), target_date="2025-12-31")
    assert report["errors"]
    assert not report["self_check"]["verification_passed"]


def test_write_word_share_data_reports_file_missing(tmp_path):
    merged_path = tmp_path / "reports" / "word_share_data.json"
    merged_path.parent.mkdir(parents=True, exist_ok=True)
    merged_path.write_text(
        json.dumps(_word_share_merged("2026-01-01"), ensure_ascii=False),
        encoding="utf-8",
    )
    report = write_word_share_data(_word_share_config(tmp_path), tmp_path, _NullLogger(), target_date="2026-01-01")
    assert report["errors"]
    assert any("词类占比 Excel 文件" in error for error in report["errors"])


# ---- 路径推导 ----

def test_get_word_share_path_derives_from_excel_path():
    project = {
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "excel": {"path": "D:\\Seafile\\竞价\\【昆明牛】\\【2026年】【昆明牛】竞价数据\\【昆明npx】2026竞价数据.xlsx"},
    }
    path = get_word_share_path(project, "2026")
    assert path.parent.name == "【2026年】【昆明牛】竞价数据"
    assert path.name == "【昆明牛】2026词类占比数据.xlsx"


# ---- GUI 命令 ----

def test_build_word_class_command_shape():
    command = build_word_class_command("C:/root", "2026-08-04", "kunming_niu")
    assert "--mode" in command
    assert command[command.index("--mode") + 1] == "run-word-class"
    assert command[command.index("--date") + 1] == "2026-08-04"
    assert command[command.index("--project") + 1] == "kunming_niu"
    assert command[-1] == "--yes"


def test_main_window_imports_build_word_class_command():
    import gui.main_window as main_window

    assert main_window.build_word_class_command is not None
    assert main_window.build_multi_word_class_command is not None


# ---- 多项目模式 ----

def test_build_multi_word_class_command_shape():
    from gui.command_builder import build_multi_word_class_command

    cmd = build_multi_word_class_command(
        "C:/root", "2026-08-03", ["kunming_niu", "shenyang_bai"]
    )
    assert cmd[cmd.index("--mode") + 1] == "run-multi"
    assert cmd[cmd.index("--task") + 1] == "word_class"
    assert cmd[cmd.index("--projects") + 1] == "kunming_niu,shenyang_bai"
    assert cmd[cmd.index("--date") + 1] == "2026-08-03"


def test_run_multi_word_class_pipeline(monkeypatch, tmp_path):
    from modules.multi_project_runner import run_multi_project_pipeline

    configs = {
        "p1": {
            "project_id": "p1",
            "project_name": "项目1",
            "excel_path": str(tmp_path / "p1" / "竞价.xlsx"),
            "baidu": {"api_profile": "p1_api"},
        },
        "p2": {
            "project_id": "p2",
            "project_name": "项目2",
            "excel_path": str(tmp_path / "p2" / "竞价.xlsx"),
            "baidu": {"api_profile": "p2_api"},
        },
    }
    fetched = []

    def fake_runtime_loader(root, project_id):
        return configs[project_id]

    def fake_fetch_search_word(config, root, logger, target_date=None, **kw):
        fetched.append(config["project_id"])
        return {
            "raw_rows": 1,
            "matched_rows": 1,
            "totals": {"click": 5, "cost": 1.0},
            "keyword_counts": {},
            "errors": [],
        }

    piped = []

    def fake_word_pipeline(config, root, logger, target_date=None, **kw):
        piped.append(config["project_id"])
        return {"passed": True, "errors": [], "failed_step": None}

    report = run_multi_project_pipeline(
        root=tmp_path,
        logger=_NullLogger(),
        project_ids=["p1", "p2"],
        task="word_class",
        target_date="2026-08-03",
        runtime_config_loader=fake_runtime_loader,
        credential_checker=lambda root, config: {"passed": True, "errors": []},
        api_fetch_search_word=fake_fetch_search_word,
        word_class_pipeline=fake_word_pipeline,
    )

    assert set(fetched) == {"p1", "p2"}  # 百度并行预取
    assert piped == ["p1", "p2"]  # 串行执行顺序
    assert report["task"] == "word_class"
    assert report["summary"]["success"] == 2
    assert report["summary"]["failed"] == 0


# ---- 人工导出模式 ----

def test_export_rows_to_word_class_conversations_maps_search_word_and_tag():
    from modules.kst_daily_aggregation import export_rows_to_word_class_conversations

    rows = [
        {"搜索关键词": "银屑病怎么治", "名片标签": "有效-三句", "访客消息数": 2},
        {"搜索关键词": "痘痘", "名片标签": None, "访客消息数": 1},
        {"搜索关键词": "银屑病", "名片标签": "", "访客消息数": 0},
        {"搜索关键词": "牛皮癣", "名片标签": "转潜-有效", "访客消息数": 3},
    ]
    convs = export_rows_to_word_class_conversations(rows)
    assert len(convs) == 3
    assert convs[0]["keyword"] == "银屑病怎么治"
    assert convs[0]["tags"] == ["有效-三句"]
    assert convs[1]["tags"] == []
    assert convs[2]["tags"] == ["转潜-有效"]
    # 无访客消息的行被排除
    assert all(conv["keyword"] != "银屑病" for conv in convs)


def test_export_rows_to_word_class_conversations_maps_dialog_keyword():
    """商务通导出文件使用'对话关键词'列(南京牛等)时,同样识别为搜索关键词。"""
    from modules.kst_daily_aggregation import export_rows_to_word_class_conversations

    rows = [
        {"对话关键词": "银屑病传染吗?", "名片标签": "智能体", "访客消息数": 3},
        {"对话关键词": "治疗牛皮癣用什么药", "名片标签": "转潜-有效", "访客消息数": 2},
        {"对话关键词": "皮肤科", "名片标签": "", "访客消息数": 0},
    ]
    convs = export_rows_to_word_class_conversations(rows)
    assert len(convs) == 2
    assert convs[0]["keyword"] == "银屑病传染吗?"
    assert convs[0]["tags"] == ["智能体"]
    assert convs[1]["keyword"] == "治疗牛皮癣用什么药"
    assert convs[1]["tags"] == ["转潜-有效"]
    # 无访客消息的行被排除
    assert all(conv["keyword"] != "皮肤科" for conv in convs)


def test_export_rows_to_word_class_conversations_accepts_unit_suffixed_messages():
    """'访客历史对话报表'的访客发送消息数带'条'单位时也能解析(宁波等)。"""
    from modules.kst_daily_aggregation import export_rows_to_word_class_conversations

    rows = [
        {"对话关键词": "牛皮癣有什么药可以吃", "名片标签": "转潜-有效-已确诊", "访客发送消息数": "1 条"},
        {"对话关键词": "银屑病怎么治", "名片标签": "智能体", "访客发送消息数": "5 条"},
        {"对话关键词": "银屑病", "名片标签": "", "访客发送消息数": "0 条"},
    ]
    convs = export_rows_to_word_class_conversations(rows)
    assert len(convs) == 2
    assert convs[0]["keyword"] == "牛皮癣有什么药可以吃"
    assert convs[0]["tags"] == ["转潜-有效-已确诊"]
    assert convs[1]["keyword"] == "银屑病怎么治"
    # '0 条' 访客消息被排除
    assert all(conv["keyword"] != "银屑病" for conv in convs)


def test_fetch_baidu_search_word_project_single_source(monkeypatch):
    from modules.baidu_report_api import fetch_baidu_search_word_project

    config = {
        "project_id": "changsha_niu",
        "project_name": "长沙牛",
        "accounts": {"A": {"baidu_names": ["A"], "kst_ids": ["111"]}},
        "baidu": {"api_profile": "changsha_niu_baidu"},
    }
    calls = []

    def fake_fetch(source_config, root, logger, target_date=None, **kw):
        calls.append(source_config)
        return {
            "raw_rows": 2,
            "matched_rows": 1,
            "totals": {"click": 2, "cost": 3.5, "impression": 10},
            "keyword_counts": {"银屑病": 1, "牛皮癣": 0},
            "errors": [],
            "diagnostics": {"api_request_count": 1, "self_heal_actions": []},
        }

    monkeypatch.setattr(
        "modules.baidu_report_api.fetch_baidu_search_word_report", fake_fetch
    )
    root = Path(tempfile.mkdtemp())
    report = fetch_baidu_search_word_project(config, root, _NullLogger(), target_date="2026-08-03")
    assert len(calls) == 1
    assert calls[0]["baidu"]["api_profile"] == "changsha_niu_baidu"
    assert report["matched_rows"] == 1
    assert report["totals"]["click"] == 2
    assert report["totals"]["cost"] == 3.5
    assert report["errors"] == []


def test_fetch_baidu_search_word_project_multi_source(monkeypatch):
    from modules.baidu_report_api import fetch_baidu_search_word_project

    config = {
        "project_id": "shenyang_bai",
        "project_name": "沈阳白",
        "baidu_sources": [
            {
                "source_id": "a",
                "source_name": "大中亚",
                "api_profile": "a_baidu",
                "accounts": [{"standard_name": "A1", "kst_ids": ["1"]}],
            },
            {
                "source_id": "b",
                "source_name": "大银康",
                "api_profile": "b_baidu",
                "accounts": [{"standard_name": "B1", "kst_ids": ["2"]}],
            },
        ],
    }

    def fake_fetch(source_config, root, logger, target_date=None, **kw):
        profile = source_config["baidu"]["api_profile"]
        if profile == "a_baidu":
            return {"search_word_rows": [
                {"queryWord": "白癜风怎么治", "click": 3, "cost": 1.0, "impression": 5}
            ], "diagnostics": {"api_request_count": 1, "self_heal_actions": []}}
        return {"search_word_rows": [
            {"queryWord": "白斑", "click": 4, "cost": 2.0, "impression": 6}
        ], "diagnostics": {"api_request_count": 1, "self_heal_actions": []}}

    monkeypatch.setattr(
        "modules.baidu_report_api.fetch_baidu_search_word_report", fake_fetch
    )
    root = Path(tempfile.mkdtemp())
    report = fetch_baidu_search_word_project(config, root, _NullLogger(), target_date="2026-08-03")
    # 沈阳白是白(白癜风)项目,只筛"白癜风"
    assert report["matched_rows"] == 1
    assert report["totals"]["click"] == 3
    assert report["totals"]["cost"] == 1.0
    assert report["errors"] == []


def test_word_class_keywords_by_project_name():
    from modules.project_config import word_class_keywords_for_name

    assert word_class_keywords_for_name("昆明牛") == ("银屑病", "牛皮癣")
    assert word_class_keywords_for_name("长沙牛") == ("银屑病", "牛皮癣")
    assert word_class_keywords_for_name("沈阳白") == ("白癜风",)
    assert word_class_keywords_for_name("青岛白") == ("白癜风",)
    assert word_class_keywords_for_name(None) == ("银屑病", "牛皮癣")


def test_fetch_baidu_search_word_project_missing_api_profile(monkeypatch):
    from modules.baidu_report_api import fetch_baidu_search_word_project

    config = {
        "project_id": "shenyang_bai",
        "project_name": "沈阳白",
        "baidu_sources": [
            {
                "source_id": "a",
                "source_name": "大中亚",
                "accounts": [{"standard_name": "A1", "kst_ids": ["1"]}],
            },
        ],
    }

    def fake_fetch(source_config, root, logger, target_date=None, **kw):
        raise AssertionError("未配置 api_profile 时不应调用搜索词接口")

    monkeypatch.setattr(
        "modules.baidu_report_api.fetch_baidu_search_word_report", fake_fetch
    )
    root = Path(tempfile.mkdtemp())
    report = fetch_baidu_search_word_project(config, root, _NullLogger(), target_date="2026-08-03")
    assert report["matched_rows"] == 0
    assert report["errors"]
    assert any("未配置百度 API 授权" in error for error in report["errors"])


def test_run_word_class_pipeline_export_branch(monkeypatch, tmp_path):
    from modules.run_pipeline import run_word_class_pipeline

    config = {
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "excel_path": str(tmp_path / "竞价.xlsx"),
        "kst": {"data_source": "export"},
    }

    def fake_find_export(root, config):
        return tmp_path / "商务通日报.xlsx"

    def fake_parse(file_path, config, root, target_date):
        return {
            "parse_report": {"passed": True, "errors": [], "row_count": 2},
            "conversations": [
                {
                    "keyword": "银屑病怎么治",
                    "tags": ["有效-三句"],
                    "visitor_messages": 2,
                },
                {
                    "keyword": "牛皮癣",
                    "tags": ["无效"],
                    "visitor_messages": 1,
                },
            ],
            "outputs": {},
        }

    def fake_write(config, root, logger, target_date=None):
        return {
            "errors": [],
            "self_check": {"verification_passed": True},
            "writes": [],
            "overwrite_summary": {"overwrite_count": 0},
            "excel_path": "",
            "backup_path": "",
        }

    def fake_search_word(config, root, logger, target_date=None, **kw):
        return {
            "raw_rows": 0,
            "matched_rows": 0,
            "totals": {"click": 0, "cost": 0.0},
            "keyword_counts": {},
            "errors": [],
        }

    monkeypatch.setattr("modules.run_pipeline.find_latest_kst_export", fake_find_export)
    monkeypatch.setattr("modules.kst_daily_parser.parse_word_class_export", fake_parse)
    monkeypatch.setattr("modules.run_pipeline.write_word_share_data", fake_write)
    monkeypatch.setattr(
        "modules.run_pipeline.fetch_baidu_search_word_project", fake_search_word
    )

    report = run_word_class_pipeline(
        config,
        tmp_path,
        _NullLogger(),
        target_date="2026-08-03",
        fetch_search_word_func=fake_search_word,
        write_func=fake_write,
    )
    assert report["passed"]
    assert report["metrics"]["有效对话"] == 1  # 仅"有效-三句"含"有效"
    assert report["metrics"]["有效转潜"] == 0


def test_run_word_class_pipeline_export_no_file(monkeypatch, tmp_path):
    from modules.run_pipeline import run_word_class_pipeline

    config = {
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "excel_path": str(tmp_path / "竞价.xlsx"),
        "kst": {"data_source": "export"},
    }

    def fake_find_export(root, config):
        return None

    def fake_write(config, root, logger, target_date=None):
        return {
            "errors": [],
            "self_check": {"verification_passed": True},
            "writes": [],
            "overwrite_summary": {"overwrite_count": 0},
            "excel_path": "",
            "backup_path": "",
        }

    def fake_search_word(config, root, logger, target_date=None, **kw):
        return {
            "raw_rows": 0,
            "matched_rows": 0,
            "totals": {"click": 0, "cost": 0.0},
            "keyword_counts": {},
            "errors": [],
        }

    monkeypatch.setattr("modules.run_pipeline.find_latest_kst_export", fake_find_export)
    monkeypatch.setattr("modules.run_pipeline.write_word_share_data", fake_write)
    monkeypatch.setattr(
        "modules.run_pipeline.fetch_baidu_search_word_project", fake_search_word
    )

    report = run_word_class_pipeline(
        config,
        tmp_path,
        _NullLogger(),
        target_date="2026-08-03",
        fetch_search_word_func=fake_search_word,
        write_func=fake_write,
    )
    assert report["passed"]
    assert report["metrics"]["有效对话"] == 0
    assert report["kst"].get("no_export_file") is True
